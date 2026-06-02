#!/usr/bin/env python3
"""Draft Generator from Template (Outlook) — ScripTree shim (Strategy A).

================================================================================
WHAT THIS PROGRAM DOES (the logic — read this to understand/rewrite the tool)
================================================================================
The user has a mailing list in an Excel workbook (``.xlsx``) — one row per
recipient, with a header row naming the columns (``Email``, ``Name``,
``Company``, ...). They also have a **subject template** and a **body template**
that contain ``{ColumnName}`` placeholders. This tool performs a *mail merge that
stops at drafts*: for every data row it creates ONE draft email in the user's
Outlook **Drafts** folder, with the recipient, subject, and body filled in from
that row. The user then reviews the drafts in Outlook and sends them by hand (or
not). Classic use: "generate 30 personalised Q3-statement drafts, let me eyeball
them before anything goes out."

================================================================================
THE HEADLINE SAFETY PROPERTY — DRAFTS ONLY, NEVER SENDS
================================================================================
This tool **creates drafts; it never sends.** That invariant is enforced in
THREE places and must survive every future edit:

  1. The generated ``.csx`` calls ONLY ``MailItem.Save()`` (which lands the
     unsent item in the Drafts folder). It contains **NO call to .Send()
     anywhere** — searching the generated script for ".Send(" must return
     nothing. Sending is forbidden; creating drafts is allowed.
  2. A **hard cap** on the number of drafts created per run. The default is 50
     and the absolute maximum is 50 — ``max_drafts`` is clamped to ``[1, 50]``
     in this shim (:func:`clamp_max`) before any rows are baked, so the ``.csx``
     can never be handed more than 50 rows. This prevents an accidental
     thousand-row blast even as drafts.
  3. The shim itself only ever *reads* the Excel file and *writes* a temp
     ``.csx``; it has no path that could send mail.

Why drafts-only matters: bulk-email is sensitive. A merge bug (wrong column,
wrong template, a stray address) is harmless if it only produces drafts a human
must still click "Send" on. It is a disaster if it auto-sends. We choose the
safe failure mode by construction.

================================================================================
THE KEY DESIGN DECISION — THE SHIM READS THE EXCEL, NOT THE .csx
================================================================================
Unlike most apps in this catalog (where the ``.csx`` does the Office work), here
the **Python shim does the Excel parsing and template rendering**, and the
``.csx`` only talks to Outlook. Rationale:

  * ``openpyxl`` is already installed in the ScripTree Python environment and is
    a far cleaner, more robust way to read ``.xlsx`` than driving Excel over COM
    from inside a Roslyn script — and it does **not** require Excel to even be
    installed or running. The only live app we need is Outlook.
  * Token rendering (``{Column}`` → cell value) is trivial string work that
    belongs in Python, not in baked C#.

So the shim:

  1. Loads the workbook with
     ``openpyxl.load_workbook(path, read_only=True, data_only=True)``
     (``read_only`` for speed/low memory on big lists; ``data_only`` so we get
     the *computed* value of any formula cell, not the formula text).
  2. Takes the **active sheet**, treats **row 1 as the headers**.
  3. Finds the recipient column by matching the ``--email-column`` header
     **case-insensitively** (error → exit 2 if that header is absent).
  4. Iterates data rows (row 2 onward). For each row it builds a
     ``{header: cell-value-as-string}`` dict, renders the subject and body by
     replacing every ``{Header}`` token (the brace-wrapped **exact** header
     text, case-sensitive) with that row's value, and **skips rows whose
     recipient cell is blank**. It stops once it has collected ``max_drafts``
     (already clamped to ≤ 50) usable rows.
  5. **Bakes** a C# array of ``(to, subject, body)`` triples into the ``.csx``
     (each field escaped as a C# string literal — the body may contain newlines,
     so the escaper handles ``\r`` and ``\n``), plus the row count.

  6. Runs ``combridge.exe outlook run-script <temp.csx> -``, parses the first
     line sentinel the ``.csx`` prints, and owns the process exit code.

If the workbook is missing or has zero usable rows, the shim **short-circuits**
with a clear message on stderr and exits 2 (it does not bother launching
Outlook/combridge for an empty merge). This is the cleaner of the two options
sketched in the spec: a guarded refusal the user can act on, rather than spinning
up COM only to report ``NOITEMS``. (The ``.csx`` still defends itself: if it is
somehow handed a zero-length array it emits ``STATUS=NOITEMS`` — see below — but
in normal operation the shim catches the empty case first.)

================================================================================
WHY A SHIM AT ALL (Strategy A — the project-wide integration pattern)
================================================================================
ScripTree is a GUI form that runs a command line: ``executable = python``, form
field values passed as argv. The Outlook work must happen inside combridge (which
owns the COM connection), and combridge's ``run-script`` runs a C# Roslyn script.
``run-script`` has **no argv channel** — a ``.csx`` sees only the plugin globals
(``olApp`` / ``olNs`` / ``olExplorer``) plus environment variables. So this shim
BAKES the data into a generated ``.csx`` (rendered from
``generate_drafts.csx.template`` by substituting ``__PLACEHOLDER__`` tokens).

================================================================================
WHY THE SENTINEL + SHIM-OWNED EXIT CODE (a hard combridge constraint)
================================================================================
combridge's ScriptHost **ignores the C# script's ``return`` value** — it exits 0
on any clean run (only a compile error = 3, an unhandled throw = 4, or a host
error = 5 produce a non-zero code). ScripTree decides success/failure from the
child process exit code. So:

  1. The ``.csx`` prints a machine-readable **sentinel** as its first stdout
     line:  ``__OLDRAFT__ STATUS=<code> [key=value ...]``  then the report.
  2. THIS shim parses that sentinel, strips it, prints the report, and
     **translates the status into the exit code** ScripTree sees.

Status → exit-code contract (see :func:`main` and the ``.csx`` template):

    STATUS=OK       -> exit 0   (drafts created; report follows. created may be 0
                                 only if every Save() failed — see failed=)
    STATUS=NOITEMS  -> exit 2   (nothing to do — zero usable rows handed to the
                                 .csx; in practice the shim catches this first)
    (combridge's own non-zero codes — 3/4/5/connect failure — pass through)

We map "nothing created" to exit 2 (a guarded no-op the caller can distinguish
from a real run) and a genuine merge to exit 0.

================================================================================
HOW COMBRIDGE IS LOCATED (portability rule)
================================================================================
combridge is NOT bundled in this project repo. Apps are authored here and
DEPLOYED into a ScripTree install that ships ``lib/combridge/combridge.exe``.
This shim finds it by walking UP the directory tree from its own location
looking for ``lib/combridge/combridge.exe`` — a relative discovery, so no
absolute path is ever baked in and the app works at whatever depth it lands.
(Not deployed here → "could not locate" + exit 1 is the correct behaviour.)
"""
from __future__ import annotations

import argparse
import os
import re
import subprocess
import sys
import tempfile
from pathlib import Path

# The first-line marker the generated .csx prints. Must match the literal used
# in generate_drafts.csx.template. Kept deliberately unlikely to collide with
# real email content.
SENTINEL = "__OLDRAFT__"

# Relative location of combridge inside a deployed ScripTree install.

# combridge plugin name for Outlook automation.
PLUGIN = "outlook"

# The hard, absolute ceiling on drafts created per run. This is the headline
# safety cap: the shim clamps the user's max_drafts into [1, MAX_DRAFTS_CAP], so
# the .csx can never be handed more than this many rows. Do NOT raise this
# without revisiting the safety story in the README + RAG.
MAX_DRAFTS_CAP = 50

# Matches a {Header} placeholder. Only used to *report* unresolved tokens for
# diagnostics; the actual rendering is exact per-header replacement (see
# render_templates) so that a literal "{" in a body that isn't a known header is
# left untouched rather than mangled.
_TOKEN_RE = re.compile(r"\{([^{}]+)\}")



def clamp_max(value: int) -> int:
    """Clamp the requested draft count into ``[1, MAX_DRAFTS_CAP]``.

    This is one of the three enforcement points of the drafts-cap safety
    invariant: no matter what the form sends, at most ``MAX_DRAFTS_CAP`` (50)
    rows are ever collected and baked. A value below 1 is meaningless (you'd
    create nothing) so it floors at 1.
    """
    if value < 1:
        return 1
    if value > MAX_DRAFTS_CAP:
        return MAX_DRAFTS_CAP
    return value


def csharp_literal(value: str) -> str:
    """Escape *value* so it is safe to drop inside a C# double-quoted string.

    The generated ``.csx`` embeds the rendered recipient/subject/body text
    directly inside ``"..."`` literals. Bodies are multi-line, so we must escape
    not just the backslash and double-quote but also ``\\r`` and ``\\n`` (a raw
    newline inside a C# regular string literal is a compile error). We escape the
    backslash FIRST so we don't double-escape the escape sequences we add after.
    A literal tab is escaped too for tidiness.
    """
    return (
        value.replace("\\", "\\\\")
        .replace('"', '\\"')
        .replace("\r", "\\r")
        .replace("\n", "\\n")
        .replace("\t", "\\t")
    )


def cell_to_str(value: object) -> str:
    """Coerce an openpyxl cell value to the string we substitute into templates.

    openpyxl returns native Python types: ``None`` for an empty cell, ``str``,
    ``int``/``float`` for numbers, ``datetime`` for dates, ``bool`` for booleans.
    A blank cell becomes the empty string (so ``{MiddleName}`` for a row with no
    middle name renders as nothing, not the text "None"). Floats that are whole
    numbers print without the trailing ``.0`` (so an ``Amount`` of ``1200`` reads
    "1200", not "1200.0"); everything else uses ``str()``.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        # Guard bool BEFORE int (bool is a subclass of int in Python).
        return "True" if value else "False"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value)


def render_templates(subject_tpl: str, body_tpl: str,
                     row: dict[str, str]) -> tuple[str, str]:
    """Render *subject_tpl* and *body_tpl* against one row's ``{header: value}``.

    Replacement is **exact, per known header**: for each header in *row* we
    replace the literal token ``"{" + header + "}"`` with that row's value. This
    is case-sensitive on the exact header text (``{Name}`` matches a header
    ``Name`` but not ``name``), and it deliberately does NOT touch braces that
    don't correspond to a real header — so a body containing, say, a JSON snippet
    or "{not a column}" is left intact rather than blanked. Longer header names
    are substituted first so a header like ``{First Name}`` isn't partially eaten
    by a shorter ``{First}`` if both exist.
    """
    def render_one(text: str) -> str:
        # Substitute longest header names first to avoid prefix collisions.
        for header in sorted(row.keys(), key=len, reverse=True):
            text = text.replace("{" + header + "}", row[header])
        return text

    return render_one(subject_tpl), render_one(body_tpl)


def read_rows(xlsx_path: Path, email_column: str, subject_tpl: str,
              body_tpl: str, max_drafts: int) -> tuple[list[tuple[str, str, str]], int]:
    """Parse the workbook and produce the baked ``(to, subject, body)`` triples.

    Returns ``(rows, scanned)`` where *rows* is the list of ready-to-bake triples
    (length ≤ *max_drafts*, already clamped to ≤ 50 by the caller) and *scanned*
    is the number of data rows examined (for the report). Raises ``ValueError``
    with a human-readable message on the two hard errors:

      * the workbook has no header row / no usable data, or
      * the *email_column* header is not present (case-insensitive).

    Logic, step by step:
      1. Open ``read_only=True, data_only=True`` and take the **active** sheet.
      2. Pull row 1 as headers (a list of strings; blank header cells are kept as
         "" placeholders so column positions line up).
      3. Locate the recipient column index by a **case-insensitive** match on the
         stripped header text against *email_column*; absent → ValueError.
      4. Walk data rows. For each, build ``{header: cell-string}`` for the
         non-blank headers, look at the recipient cell, **skip the row if that
         cell is blank** (after stripping), else render subject+body and append
         the triple. Stop at *max_drafts* collected rows.
    """
    from openpyxl import load_workbook  # imported here so a missing-file error
                                        # path doesn't pay the import cost

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("The workbook has no active worksheet.")

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("The worksheet is empty (no header row).")

        headers = [cell_to_str(h).strip() for h in header_row]
        if not any(headers):
            raise ValueError("Row 1 has no column headers.")

        # Case-insensitive match for the recipient column.
        target = email_column.strip().lower()
        email_idx = None
        for i, h in enumerate(headers):
            if h.lower() == target:
                email_idx = i
                break
        if email_idx is None:
            available = ", ".join(h for h in headers if h) or "(none)"
            raise ValueError(
                f"No column header matches the recipient column "
                f"'{email_column}' (case-insensitive). "
                f"Available headers: {available}.")

        out: list[tuple[str, str, str]] = []
        scanned = 0
        for raw in rows_iter:
            scanned += 1
            # Build the {header: value} dict for this row (skip blank headers).
            row: dict[str, str] = {}
            recipient = ""
            for i, header in enumerate(headers):
                cell = raw[i] if i < len(raw) else None
                text = cell_to_str(cell)
                if header:
                    row[header] = text
                if i == email_idx:
                    recipient = text.strip()

            if not recipient:
                continue  # skip rows with a blank recipient cell

            subject, body = render_templates(subject_tpl, body_tpl, row)
            out.append((recipient, subject, body))
            if len(out) >= max_drafts:
                break

        return out, scanned
    finally:
        wb.close()


def bake_rows(triples: list[tuple[str, str, str]]) -> str:
    """Render the C# array initialiser for the baked ``(to, subject, body)`` rows.

    Produces a sequence of ``new string[] { "to", "subj", "body" },`` lines, each
    field run through :func:`csharp_literal`. The template wraps this in a
    ``var ROWS = new List<string[]> { ... };`` so the ``.csx`` just iterates it.
    An empty list renders to nothing (the count token ``__ROW_COUNT__`` is 0 and
    the ``.csx`` reports NOITEMS).
    """
    parts = []
    for to, subj, body in triples:
        parts.append(
            '    new string[] { "%s", "%s", "%s" },'
            % (csharp_literal(to), csharp_literal(subj), csharp_literal(body)))
    return "\n".join(parts)


def render_csx(template: str, *, triples: list[tuple[str, str, str]],
               mailing_list: str, output_format: str) -> str:
    """Fill the ``.csx`` template with the baked rows and metadata.

    Every ``__TOKEN__`` in the template must be replaced here — a leftover token
    would be a compile error, which is why the offline render-check greps the
    output for ``__[A-Z_]+__`` and expects only the ``__OLDRAFT__`` sentinel to
    survive. ``__ROW_COUNT__`` is baked as a bare integer (no quotes); the
    string metadata goes through :func:`csharp_literal`.
    """
    return (
        template.replace("__BAKED_ROWS__", bake_rows(triples))
        .replace("__ROW_COUNT__", str(len(triples)))
        .replace("__MAILING_LIST__", csharp_literal(mailing_list))
        .replace("__OUTPUT_FORMAT__", csharp_literal(output_format))
    )


def main() -> int:
    """Parse argv, read the Excel, render the ``.csx``, run combridge, translate.

    Returns the process exit code ScripTree should see (see the status→exit
    contract in the module docstring).
    """
    parser = argparse.ArgumentParser(
        description="Create one DRAFT Outlook email per row of an Excel mailing "
                    "list, filling {Column} placeholders in the subject/body "
                    "templates. NEVER sends — drafts only, capped at 50/run.")
    parser.add_argument(
        "--mailing-list", dest="mailing_list", required=True,
        help="Path to the .xlsx mailing list. Row 1 = headers; one recipient "
             "per data row.")
    parser.add_argument(
        "--email-column", dest="email_column", default="Email",
        help="Header text of the column holding recipient addresses "
             "(case-insensitive). Default 'Email'.")
    parser.add_argument(
        "--subject-template", dest="subject_template", default="",
        help="Subject line; may contain {Column} placeholders, e.g. "
             "'Your Q3 statement, {Name}'.")
    parser.add_argument(
        "--body-template", dest="body_template", required=True,
        help="Body text (multi-line); may contain {Column} placeholders.")
    parser.add_argument(
        "--max-drafts", dest="max_drafts", type=int, default=MAX_DRAFTS_CAP,
        help=f"Maximum drafts to create this run. Clamped to "
             f"[1, {MAX_DRAFTS_CAP}].")
    parser.add_argument(
        "--output-format", dest="output_format", default="markdown",
        choices=["markdown", "text"])
    args = parser.parse_args()

    mailing_list = args.mailing_list.strip()
    if not mailing_list:
        print("ERROR: --mailing-list must not be empty.", file=sys.stderr)
        return 2

    xlsx_path = Path(mailing_list)
    if not xlsx_path.is_file():
        print(f"ERROR: mailing list not found: {xlsx_path}", file=sys.stderr)
        return 2

    # Enforce the hard cap BEFORE reading rows (cannot collect more than 50).
    max_drafts = clamp_max(args.max_drafts)

    try:
        triples, scanned = read_rows(
            xlsx_path,
            email_column=args.email_column,
            subject_tpl=args.subject_template,
            body_tpl=args.body_template,
            max_drafts=max_drafts,
        )
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    except Exception as exc:  # malformed workbook, openpyxl read error, etc.
        print(f"ERROR: could not read the mailing list: {exc}", file=sys.stderr)
        return 2

    # Short-circuit the empty merge: don't spin up Outlook/combridge for zero
    # usable rows. (The .csx still defends itself with NOITEMS if ever handed an
    # empty array, but this is the cleaner first line of defence.)
    if not triples:
        print(
            "ERROR: no usable rows in the mailing list — every data row had a "
            "blank recipient cell, or the sheet had only a header row. Nothing "
            "to draft.",
            file=sys.stderr,
        )
        return 2

    here = Path(__file__).resolve().parent
    template_path = here / "generate_drafts.csx.template"
    if not template_path.is_file():
        print(f"ERROR: template not found: {template_path}", file=sys.stderr)
        return 1

    template = template_path.read_text(encoding="utf-8")
    csx = render_csx(
        template,
        triples=triples,
        mailing_list=mailing_list,
        output_format=args.output_format,
    )

    # Write the generated script to a temp .csx, run it, always clean up.
    # delete=False + manual unlink in finally is required on Windows so the
    # child process can open the file while we hold no lock on it.
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csx", prefix="generate_drafts_",
        delete=False, encoding="utf-8")
    try:
        tmp.write(csx)
        tmp.close()
        proc = subprocess.run(
            ["combridge.exe", PLUGIN, "run-script", tmp.name, "-"],
            capture_output=True, text=True,
            encoding="utf-8", errors="replace",
        )
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass

    if proc.stderr:
        sys.stderr.write(proc.stderr)

    # combridge failed before/while running the script (compile=3, throw=4,
    # host=5, or could not connect to Outlook): surface output, propagate code.
    if proc.returncode != 0:
        sys.stdout.write(proc.stdout)
        return proc.returncode

    # Clean run: parse the sentinel first line, print the rest as the report.
    lines = proc.stdout.splitlines()
    status = "OK"
    body_start = 0
    if lines and lines[0].startswith(SENTINEL):
        fields = lines[0].split()
        for f in fields[1:]:
            if f.startswith("STATUS="):
                status = f[len("STATUS="):]
        body_start = 1

    body = "\n".join(lines[body_start:])
    if body:
        print(body)

    # Any guarded no-op (NOITEMS) maps to exit 2; a real merge is 0.
    return 0 if status == "OK" else 2


if __name__ == "__main__":
    raise SystemExit(main())
