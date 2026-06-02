#!/usr/bin/env python3
"""Generate a workbook with a REAL external link for the Link Auditor.

Run:  python make_example.py
Produces (next to this script):
    external_source.xlsx   — the workbook the link points AT
    dashboard.xlsx         — has an external link to external_source.xlsx

WHY THIS IS NON-TRIVIAL
-----------------------
openpyxl can write cell formulas but does NOT create the OOXML
*external-link* parts that make Excel treat a workbook as having a
linked source (and that make ``Workbook.LinkSources(xlExcelLinks)`` —
what the auditor calls — return the source path). A real external link
needs five coordinated pieces inside the .xlsx zip:

  1. xl/externalLinks/externalLink1.xml         — the link + cached values
  2. xl/externalLinks/_rels/externalLink1.xml.rels — target path (External)
  3. xl/workbook.xml  <externalReferences>       — declares the reference
  4. xl/_rels/workbook.xml.rels                  — workbook → externalLink
  5. [Content_Types].xml  Override               — content type of part 1

So we build dashboard.xlsx with openpyxl (including a formula that USES
the link, ``=[1]Data!A1``), then re-open the zip and inject parts 1–5.
The link target is written as an ABSOLUTE ``file:///`` path to
``external_source.xlsx`` **in this same folder**, resolved at generation
time (we can't know the path until you run this on your machine).

This is "best-effort real": it produces a genuinely linked workbook.
Like every app in this catalog it is pending live verification against a
real Excel — if a given Excel build declines to surface the injected
link, fall back to the scenario in README.md.
"""
import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).resolve().parent

NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
EXTLINK_TYPE = NS_REL + "/externalLink"
EXTPATH_TYPE = NS_REL + "/externalLinkPath"

EXTERNAL_LINK_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <externalBook r:id="rId1">
    <sheetNames><sheetName val="Data"/></sheetNames>
    <sheetDataSet>
      <sheetData sheetId="0">
        <row r="1"><cell r="A1"><v>42</v></cell></row>
      </sheetData>
    </sheetDataSet>
  </externalBook>
</externalLink>"""


def external_rels_xml(target_uri: str) -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'<Relationship Id="rId1" Type="{EXTPATH_TYPE}" '
        f'Target="{target_uri}" TargetMode="External"/>'
        "</Relationships>"
    )


def build_source() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 42
    ws["A2"] = "the answer"
    wb.save(str(HERE / "external_source.xlsx"))
    print(f"wrote {HERE / 'external_source.xlsx'}")


def build_dashboard_base() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "Linked value from external_source.xlsx:"
    # A formula that USES external reference [1] (the first/only externalReference).
    ws["B1"] = "=[1]Data!A1"
    path = HERE / "dashboard.xlsx"
    wb.save(str(path))
    return path


def inject_external_link(xlsx_path: Path, target_uri: str) -> None:
    tmp = xlsx_path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = zin.namelist()
        data = {n: zin.read(n) for n in names}

    # 5. [Content_Types].xml — add the externalLink Override.
    ct = data["[Content_Types].xml"].decode("utf-8")
    override = (
        '<Override PartName="/xl/externalLinks/externalLink1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.'
        'spreadsheetml.externalLink+xml"/>'
    )
    ct = ct.replace("</Types>", override + "</Types>")
    data["[Content_Types].xml"] = ct.encode("utf-8")

    # 4. xl/_rels/workbook.xml.rels — workbook → externalLink relationship.
    rels = data["xl/_rels/workbook.xml.rels"].decode("utf-8")
    # pick an Id not already used
    used = set(re.findall(r'Id="(rId\d+)"', rels))
    n = 1
    while f"rId{n}" in used:
        n += 1
    ext_rid = f"rId{n}"
    rel = (
        f'<Relationship Id="{ext_rid}" Type="{EXTLINK_TYPE}" '
        'Target="externalLinks/externalLink1.xml"/>'
    )
    rels = rels.replace("</Relationships>", rel + "</Relationships>")
    data["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")

    # 3. xl/workbook.xml — declare <externalReferences> right after </sheets>
    #    (CT_Workbook schema order: sheets, functionGroups, externalReferences…).
    wbxml = data["xl/workbook.xml"].decode("utf-8")
    # openpyxl declares xmlns:r locally on <sheet>, not on the root <workbook>,
    # so we must declare it on our injected element too or the r:id is undefined.
    ext_refs = (
        "<externalReferences>"
        f'<externalReference xmlns:r="{NS_REL}" r:id="{ext_rid}"/>'
        "</externalReferences>"
    )
    wbxml = wbxml.replace("</sheets>", "</sheets>" + ext_refs, 1)
    data["xl/workbook.xml"] = wbxml.encode("utf-8")

    # 1 & 2. the externalLink part and its rels.
    data["xl/externalLinks/externalLink1.xml"] = EXTERNAL_LINK_XML.encode("utf-8")
    data["xl/externalLinks/_rels/externalLink1.xml.rels"] = external_rels_xml(
        target_uri
    ).encode("utf-8")

    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, blob in data.items():
            zout.writestr(name, blob)
    shutil.move(str(tmp), str(xlsx_path))


def main() -> None:
    build_source()
    dash = build_dashboard_base()
    # Absolute file:/// URI to the sibling source workbook on THIS machine.
    target_uri = (HERE / "external_source.xlsx").as_uri()
    inject_external_link(dash, target_uri)
    print(f"wrote {dash}  (external link -> {target_uri})")


if __name__ == "__main__":
    main()
