#!/usr/bin/env python3
"""Generate the sample formula workbook for the Formula-to-Value Freezer.

Run:  python make_example.py
Produces:  sample_formulas.xlsx  (next to this script)

WHAT THE SAMPLE CONTAINS (and why)
----------------------------------
A workbook with TWO worksheets full of real formulas, so that after freezing
you can confirm every formula cell became a static value:

  Sheet "Sales"
    A header row: Item | Qty | Price | Total
    4 data rows; the Total column is a per-row formula  =B*C  (Qty * Price).
    A summary block underneath:
       "Units"     =SUM(Qty range)        (a SUM)
       "Avg price" =AVERAGE(Price range)  (an AVERAGE)
       "Revenue"   =SUM(Total range)      (a SUM over the formula column)

  Sheet "Summary"
    Pulls figures from Sales via CROSS-SHEET references:
       "Total revenue"   =Sales!<Revenue cell>
       "Total units"     =Sales!<Units cell>
       "Average price"   =Sales!<Avg price cell>
       "Revenue / unit"  =<Total revenue> / <Total units>   (a formula on formulas)

After running the Freezer (scope = All worksheets, Work on a copy = ON), the
sibling sample_formulas_Frozen.xlsx has the SAME displayed numbers but every
one of those cells now holds a constant — no formula. The original
sample_formulas.xlsx is left untouched.

ASCII-only output in print() so this runs cleanly under any console encoding.
"""
from pathlib import Path

from openpyxl import Workbook


def main() -> None:
    wb = Workbook()

    # ----- Sheet 1: Sales (per-row formulas + summary aggregates) -----------
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Item", "Qty", "Price", "Total"])
    data = [
        ("Widget", 120, 4.50),
        ("Gadget", 80, 12.00),
        ("Gizmo", 60, 7.25),
        ("Doohickey", 45, 19.99),
    ]
    first_data_row = 2
    for i, (item, qty, price) in enumerate(data):
        r = first_data_row + i
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=qty)
        ws.cell(row=r, column=3, value=price)
        # Per-row Total formula: Qty * Price
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
    last_data_row = first_data_row + len(data) - 1

    # Summary block under the table.
    sums_row = last_data_row + 2          # leave one blank row
    ws.cell(row=sums_row, column=1, value="Units")
    units_cell = f"Sales!B{sums_row + 1}"  # placeholder; we set the value cell next
    ws.cell(row=sums_row, column=2,
            value=f"=SUM(B{first_data_row}:B{last_data_row})")
    ws.cell(row=sums_row + 1, column=1, value="Avg price")
    ws.cell(row=sums_row + 1, column=2,
            value=f"=AVERAGE(C{first_data_row}:C{last_data_row})")
    ws.cell(row=sums_row + 2, column=1, value="Revenue")
    ws.cell(row=sums_row + 2, column=2,
            value=f"=SUM(D{first_data_row}:D{last_data_row})")

    units_addr = f"Sales!B{sums_row}"
    avg_addr = f"Sales!B{sums_row + 1}"
    revenue_addr = f"Sales!B{sums_row + 2}"

    # ----- Sheet 2: Summary (cross-sheet references) ------------------------
    s2 = wb.create_sheet("Summary")
    s2.append(["Metric", "Value"])
    s2.cell(row=2, column=1, value="Total revenue")
    s2.cell(row=2, column=2, value=f"={revenue_addr}")
    s2.cell(row=3, column=1, value="Total units")
    s2.cell(row=3, column=2, value=f"={units_addr}")
    s2.cell(row=4, column=1, value="Average price")
    s2.cell(row=4, column=2, value=f"={avg_addr}")
    # A formula built on other formulas (revenue per unit).
    s2.cell(row=5, column=1, value="Revenue / unit")
    s2.cell(row=5, column=2, value="=B2/B3")

    out = Path(__file__).resolve().parent / "sample_formulas.xlsx"
    wb.save(str(out))
    print("wrote " + str(out))
    print("Sheets: Sales (per-row =B*C + SUM/AVERAGE) and Summary (cross-sheet refs).")


if __name__ == "__main__":
    main()
